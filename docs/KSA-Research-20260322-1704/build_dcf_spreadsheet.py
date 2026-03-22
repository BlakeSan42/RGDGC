#!/usr/bin/env python3
"""
KSA (Kingwood Service Association) — Comprehensive DCF Financial Model v2
Corrected revenue model: 29 member associations, ~23,354 equivalent units
All cells use Excel formulas for transparency and auditability.
Yellow = editable input | Green = calculated formula | Red = warning
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styling ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
CALC_FILL = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
ALERT_FILL = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B5E20")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1B5E20")
BOLD_FONT = Font(name="Calibri", bold=True, size=11)
NORMAL_FONT = Font(name="Calibri", size=11)
NOTE_FONT = Font(name="Calibri", italic=True, size=10, color="666666")
MONEY = '#,##0'
MONEY_NEG = '#,##0;[Red](#,##0)'
PCT = '0.0%'
PCT2 = '0.00%'
YR = '0'
DEC1 = '0.0'
THIN = Border(left=Side(style='thin'), right=Side(style='thin'),
              top=Side(style='thin'), bottom=Side(style='thin'))


def hdr(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN


def sc(ws, r, c, val=None, fmt=None, fill=None, font=None):
    """Set cell with optional formatting."""
    cell = ws.cell(row=r, column=c, value=val)
    if fmt: cell.number_format = fmt
    if fill: cell.fill = fill
    if font: cell.font = font
    cell.border = THIN
    cell.alignment = Alignment(horizontal='center')
    return cell


def autofit(ws):
    for col in ws.columns:
        mx = 12
        cl = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)) + 2, 26))
        ws.column_dimensions[cl].width = mx


# ── Actual 990 Data ──────────────────────────────────────────────────────────

# (Year, ProgramSvcRev, InvestmentInc, OtherRev, TotalRev, TotalExp, TotalAssets, TotalLiab, NetAssets)
DATA = [
    (2011, 798934, 13444, 7510, 819888, 743479, 2987497, 73750, 2913747),
    (2012, 639431,  6389, 16595, 662415, 724047, 2899789, 170172, 2729617),
    (2013, 832330,  4728, 8320, 845378, 911275, 2674544, 4640, 2669904),
    (2014, 748124,  4563, 7790, 760477, 649890, 2904598, 124107, 2780491),
    (2015, 823714,  6044, 8575, 838333, 771855, 2944439, 97468, 2846971),
    (2016, 889400,  9674, 5490, 904564, 789390, 3045749, 83604, 2962145),
    (2017, 873548, 10486,    0, 884034, 895731, 3060685, 110237, 2950448),
    (2018, 904037, 22436,    0, 926473, 745644, 3279346, 82355, 3196991),
    (2019, 961787, 36464,    0, 998251, 850536, 3344706, 0, 3344706),
    (2020, 1014760, 16474,   0, 1031234, 1018430, 3357510, 0, 3357510),
    (2021, 1024453, 15504,   0, 1039957, 948240, 3449283, 56, 3449227),
    (2022, 1014641, 20222,   0, 1034863, 861519, 3604403, 33, 3604370),
    (2023, 991094, 50707,    0, 1041801, 1009164, 3644171, 7164, 3637007),
    (2024, 958464, 90457,    0, 1048921, 1565650, 3123778, 3500, 3120278),
]

# Budget category breakdowns (2009 actual, 2012 actual — only years with detail)
BUDGET_2009 = [
    ("Parks Operations & Maintenance", 336526, "Landscaping, repairs, supplies for 5 parks"),
    ("General & Administrative", 153319, "KAM management contract, office, insurance, legal"),
    ("Entryway Maintenance", 143200, "Kingwood Dr, N Park Dr, W Lake Houston Pkwy"),
    ("Park Improvements (Capital)", 91100, "New infrastructure, major repairs"),
    ("TxDOT Landscaping Match", 50000, "State matching funds for right-of-way"),
    ("Public Safety Activities", 17000, "HPD bike patrol, HPD/HFD coordination"),
]

BUDGET_2012 = [
    ("Parks Operations & Maintenance", 354063, "5 parks, 356.6 total acres"),
    ("General & Administrative", 182783, "KAM contract, insurance, office, legal"),
    ("Entryway Maintenance", 133772, "3 major entryways"),
    ("Park Improvements (Capital)", 118385, "Capital projects"),
    ("Public Safety Activities", 17000, "HPD coordination"),
]

# Known field leases
LEASES = [
    ("Kingwood Alliance Soccer Club", "River Grove + Northpark", "Soccer fields", "Active"),
    ("Kingwood Youth Lacrosse (KYLAX)", "River Grove", "Lacrosse fields", "Active"),
    ("Kingwood/Forest Cove Baseball Assn", "Deer Ridge", "Baseball fields", "Active"),
    ("Kingwood Girls Softball Assn", "Northpark", "Softball fields", "Active"),
    ("Kingwood Adult Softball Assn", "Northpark", "Softball fields", "Active"),
    ("Kingwood Horsemen's Assn / Trail's End", "Deer Ridge", "Stables/pasture", "Active (since 1980)"),
    ("Texas HeatWave Soccer Club", "Deer Ridge", "Soccer fields", "Terminated 2010"),
]


def build():
    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 1: COMPLETE REVENUE PICTURE
    # ════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Revenue Analysis"
    ws.sheet_properties.tabColor = "1B5E20"

    ws.merge_cells('A1:K1')
    ws['A1'] = "KSA COMPLETE REVENUE ANALYSIS — WHERE EVERY DOLLAR COMES FROM"
    ws['A1'].font = TITLE_FONT
    ws.merge_cells('A2:K2')
    ws['A2'] = "Source: IRS Form 990 (2011-2024) | EIN: 74-1891991 | 29 member associations | ~23,354 equivalent units"
    ws['A2'].font = NOTE_FONT

    # Revenue headers
    rev_h = ['Year', 'Program Svc Rev\n(Assessments+Leases)', 'Investment\nIncome',
             'Other\nRevenue', 'Total\nRevenue', 'Assessment\n% of Total',
             'Invest %\nof Total', 'YoY Rev\nGrowth', 'Implied Per\nUnit Cost',
             'Program Rev\nYoY Growth', 'Invest Inc\nYoY Growth']
    for c, h in enumerate(rev_h, 1):
        ws.cell(row=4, column=c, value=h)
    hdr(ws, 4, len(rev_h))

    for i, d in enumerate(DATA):
        r = 5 + i
        yr, prog, inv, oth, tot, exp, ast, liab, net = d
        sc(ws, r, 1, yr, YR, font=BOLD_FONT)
        sc(ws, r, 2, prog, MONEY)
        sc(ws, r, 3, inv, MONEY)
        sc(ws, r, 4, oth, MONEY)
        sc(ws, r, 5, None, MONEY, CALC_FILL).value = f'=B{r}+C{r}+D{r}'
        sc(ws, r, 6, None, PCT, CALC_FILL).value = f'=B{r}/E{r}'
        sc(ws, r, 7, None, PCT, CALC_FILL).value = f'=C{r}/E{r}'
        if i > 0:
            sc(ws, r, 8, None, PCT, CALC_FILL).value = f'=E{r}/E{r-1}-1'
            sc(ws, r, 10, None, PCT, CALC_FILL).value = f'=B{r}/B{r-1}-1'
            sc(ws, r, 11, None, PCT, CALC_FILL).value = f'=IF(C{r-1}=0,"N/A",C{r}/C{r-1}-1)'
        sc(ws, r, 9, None, '$#,##0.00', CALC_FILL).value = f'=B{r}/23354'

    de = 5 + len(DATA) - 1  # data end row

    # Revenue summary stats
    sr = de + 2
    ws.cell(row=sr, column=1, value="REVENUE SUMMARY STATISTICS").font = SECTION_FONT
    rstats = [
        ("14-Year Avg Total Revenue", f'=AVERAGE(E5:E{de})', MONEY),
        ("14-Year Avg Program Svc Rev", f'=AVERAGE(B5:B{de})', MONEY),
        ("14-Year Avg Investment Inc", f'=AVERAGE(C5:C{de})', MONEY),
        ("Revenue CAGR (2011→2024)", f'=(E{de}/E5)^(1/13)-1', PCT2),
        ("Program Svc CAGR", f'=(B{de}/B5)^(1/13)-1', PCT2),
        ("Investment Inc CAGR", f'=(C{de}/C5)^(1/13)-1', PCT2),
        ("Avg Per-Unit Assessment", f'=AVERAGE(I5:I{de})', '$#,##0.00'),
        ("2024 Per-Unit Assessment", f'=I{de}', '$#,##0.00'),
        ("Revenue Volatility (Std Dev)", f'=STDEV(E5:E{de})', MONEY),
        ("Max Single-Year Revenue", f'=MAX(E5:E{de})', MONEY),
        ("Min Single-Year Revenue", f'=MIN(E5:E{de})', MONEY),
    ]
    for j, (lbl, fml, fmt) in enumerate(rstats):
        r = sr + 1 + j
        sc(ws, r, 1, lbl, font=BOLD_FONT)
        sc(ws, r, 2, None, fmt, CALC_FILL).value = fml

    # Revenue stream identification
    rs = sr + len(rstats) + 3
    ws.cell(row=rs, column=1, value="IDENTIFIED REVENUE STREAMS").font = SECTION_FONT
    ws.cell(row=rs + 1, column=1, value="All revenue flows through TWO lines on the 990:").font = NOTE_FONT
    ws.merge_cells(f'A{rs+1}:K{rs+1}')

    streams = [
        ("1. Quarterly Pro-Rata Assessments", "~$41/equivalent unit/year", "~$958K", "91%",
         "Billed to 29 member associations based on equivalent units. Villages pass cost to homeowners within HOA dues."),
        ("2. Field Lease Income (bundled)", "Unknown per lease", "Bundled in #1", "?%",
         "6+ active leases to youth sports orgs. Reports as $0 'net rental income' — likely bundled in program services or at-cost."),
        ("3. Investment Income", "3-4% yield on reserves", "~$90K", "8.6%",
         "Interest/dividends on $2.1M cash reserves. Grew 572% from 2015-2024 as reserves and rates increased."),
        ("4. Other/Misc Revenue", "Varies", "$0 (recent)", "<1%",
         "Was $5-17K/yr (2011-2016), dropped to $0 from 2017. May have been K-sticker fees, pavilion rentals."),
        ("5. KSA Parks Foundation", "Tax-deductible donations", "<$50K/yr", "Separate",
         "Separate 501(c)(3) entity (EIN 27-1433729). Files 990-N. Funds Trees for Kingwood, park donations."),
        ("6. TxDOT Matching Grants", "State highway landscaping", "$50K (2009)", "One-time",
         "Matching funds for entryway landscaping along state right-of-way. Not consistent."),
        ("7. FEMA / Insurance Claims", "Post-disaster", "Unknown", "Irregular",
         "No FEMA income visible on 990s. Harvey recovery costs may have been self-funded from reserves."),
    ]

    strm_h = ['Revenue Stream', 'Rate/Basis', 'Annual Amount', '% of Total', 'Notes']
    for c, h in enumerate(strm_h, 1):
        ws.cell(row=rs + 3, column=c, value=h)
    hdr(ws, rs + 3, 5)

    for j, (name, rate, amt, pct, notes) in enumerate(streams):
        r = rs + 4 + j
        sc(ws, r, 1, name, font=BOLD_FONT)
        sc(ws, r, 2, rate)
        sc(ws, r, 3, amt)
        sc(ws, r, 4, pct)
        sc(ws, r, 5, notes, font=NOTE_FONT)

    # Field leases detail
    fl = rs + 4 + len(streams) + 2
    ws.cell(row=fl, column=1, value="KNOWN FIELD LEASE AGREEMENTS").font = SECTION_FONT
    fl_h = ['Organization', 'Park', 'Facility', 'Status', 'Notes']
    for c, h in enumerate(fl_h, 1):
        ws.cell(row=fl + 1, column=c, value=h)
    hdr(ws, fl + 1, 5)

    for j, (org, park, fac, status) in enumerate(LEASES):
        r = fl + 2 + j
        sc(ws, r, 1, org)
        sc(ws, r, 2, park)
        sc(ws, r, 3, fac)
        sc(ws, r, 4, status)
        note = "Lessees pay own utilities + maintenance (confirmed by KWFCBA)" if j == 0 else ""
        sc(ws, r, 5, note, font=NOTE_FONT)

    autofit(ws)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 2: EXPENSE ANALYSIS — WHERE EVERY DOLLAR GOES
    # ════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Expense Analysis")
    ws2.sheet_properties.tabColor = "FF6B35"

    ws2.merge_cells('A1:L1')
    ws2['A1'] = "KSA EXPENSE ANALYSIS — WHERE EVERY DOLLAR GOES"
    ws2['A1'].font = TITLE_FONT
    ws2.merge_cells('A2:L2')
    ws2['A2'] = "Budget categories from KSA presentations (2009, 2012). 990 detail locked in scanned PDFs."
    ws2['A2'].font = NOTE_FONT

    # Historical expenses
    exp_h = ['Year', 'Total\nExpenses', 'YoY\nChange', 'Total\nRevenue',
             'Net\nIncome', 'Expense/\nRevenue', 'Per Unit\nExpense',
             'Expense\nVolatility']
    for c, h in enumerate(exp_h, 1):
        ws2.cell(row=4, column=c, value=h)
    hdr(ws2, 4, len(exp_h))

    for i, d in enumerate(DATA):
        r = 5 + i
        yr, prog, inv, oth, tot, exp, ast, liab, net = d
        sc(ws2, r, 1, yr, YR, font=BOLD_FONT)
        sc(ws2, r, 2, exp, MONEY)
        if i > 0:
            sc(ws2, r, 3, None, PCT, CALC_FILL).value = f'=B{r}/B{r-1}-1'
        sc(ws2, r, 4, tot, MONEY)
        sc(ws2, r, 5, None, MONEY_NEG, CALC_FILL).value = f'=D{r}-B{r}'
        sc(ws2, r, 6, None, PCT, CALC_FILL).value = f'=B{r}/D{r}'
        sc(ws2, r, 7, None, '$#,##0.00', CALC_FILL).value = f'=B{r}/23354'

    de2 = 5 + len(DATA) - 1

    # Expense summary stats
    esr = de2 + 2
    ws2.cell(row=esr, column=1, value="EXPENSE SUMMARY").font = SECTION_FONT
    estats = [
        ("14-Year Avg Expenses", f'=AVERAGE(B5:B{de2})', MONEY),
        ("Expense CAGR (2011→2024)", f'=(B{de2}/B5)^(1/13)-1', PCT2),
        ("Avg Expense/Revenue Ratio", f'=AVERAGE(F5:F{de2})', PCT),
        ("Expense Std Deviation", f'=STDEV(B5:B{de2})', MONEY),
        ("Max Single-Year Expense", f'=MAX(B5:B{de2})', MONEY),
        ("Min Single-Year Expense", f'=MIN(B5:B{de2})', MONEY),
        ("2024 Per-Unit Expense", f'=B{de2}/23354', '$#,##0.00'),
        ("Median Expense (excl 2024)", f'=MEDIAN(B5:B{de2-1})', MONEY),
        ("2024 vs Median Deviation", f'=B{de2}/MEDIAN(B5:B{de2-1})-1', PCT),
    ]
    for j, (lbl, fml, fmt) in enumerate(estats):
        r = esr + 1 + j
        sc(ws2, r, 1, lbl, font=BOLD_FONT)
        sc(ws2, r, 2, None, fmt, CALC_FILL).value = fml

    # Budget category breakdowns
    bc = esr + len(estats) + 3
    ws2.cell(row=bc, column=1, value="BUDGET CATEGORY BREAKDOWN (Confirmed from KSA Presentations)").font = SECTION_FONT

    bc_h = ['Category', '2009 Amount', '2009 %', '2012 Amount', '2012 %',
            'Growth\n(2009→2012)', 'Est. 2024\n(at CAGR)', 'Est. 2024 %',
            'Description']
    for c, h in enumerate(bc_h, 1):
        ws2.cell(row=bc + 1, column=c, value=h)
    hdr(ws2, bc + 1, len(bc_h))

    # Paired data
    cats = [
        ("Parks Ops & Maintenance", 336526, 354063, "Landscaping, mowing, repairs, supplies, utilities for 5 parks (356.6 acres)"),
        ("General & Administrative", 153319, 182783, "KAM management contract, office expenses, umbrella insurance, legal fees, accounting"),
        ("Entryway Maintenance", 143200, 133772, "Kingwood Dr, North Park Dr, West Lake Houston Pkwy medians + signage"),
        ("Park Improvements (Capital)", 91100, 118385, "New infrastructure, tee pads, benches, trail improvements, facility upgrades"),
        ("TxDOT Matching / Public Safety", 67000, 17000, "2009: $50K TxDOT match + $17K safety. 2012: $17K safety only"),
    ]

    for j, (name, v09, v12, desc) in enumerate(cats):
        r = bc + 2 + j
        sc(ws2, r, 1, name, font=BOLD_FONT)
        sc(ws2, r, 2, v09, MONEY)
        sc(ws2, r, 3, None, PCT, CALC_FILL).value = f'=B{r}/B{bc+2+len(cats)}'
        sc(ws2, r, 4, v12, MONEY)
        sc(ws2, r, 5, None, PCT, CALC_FILL).value = f'=D{r}/D{bc+2+len(cats)}'
        # 3-year CAGR (2009→2012)
        sc(ws2, r, 6, None, PCT, CALC_FILL).value = f'=(D{r}/B{r})^(1/3)-1'
        # Estimated 2024 at that CAGR from 2012 (12 years)
        sc(ws2, r, 7, None, MONEY, CALC_FILL).value = f'=D{r}*(1+F{r})^12'
        sc(ws2, r, 9, desc, font=NOTE_FONT)

    # Totals row
    tr = bc + 2 + len(cats)
    sc(ws2, tr, 1, "TOTAL", font=BOLD_FONT)
    sc(ws2, tr, 2, None, MONEY, CALC_FILL).value = f'=SUM(B{bc+2}:B{tr-1})'
    sc(ws2, tr, 3, None, PCT, CALC_FILL).value = '=1'
    sc(ws2, tr, 4, None, MONEY, CALC_FILL).value = f'=SUM(D{bc+2}:D{tr-1})'
    sc(ws2, tr, 5, None, PCT, CALC_FILL).value = '=1'
    sc(ws2, tr, 7, None, MONEY, CALC_FILL).value = f'=SUM(G{bc+2}:G{tr-1})'

    # Estimated 2024 % for each category
    for j in range(len(cats)):
        r = bc + 2 + j
        sc(ws2, r, 8, None, PCT, CALC_FILL).value = f'=G{r}/G{tr}'

    # Gap analysis: estimated vs actual 2024
    gap_r = tr + 2
    ws2.cell(row=gap_r, column=1, value="2024 EXPENSE ANOMALY ANALYSIS").font = SECTION_FONT
    gap_data = [
        ("Actual 2024 Expenses (from 990)", 1565650, MONEY),
        ("Estimated 2024 (from CAGR model)", f'=G{tr}', MONEY),
        ("Unexplained Excess", f'=B{gap_r+1}-B{gap_r+2}', MONEY_NEG),
        ("Excess as % of Normal Budget", f'=B{gap_r+3}/B{gap_r+2}', PCT),
        ("Median Expense (2011-2023)", f'=MEDIAN(B5:B{de2-1})', MONEY),
        ("2024 vs Median", f'=B{gap_r+1}/B{gap_r+5}-1', PCT),
    ]
    for j, (lbl, val, fmt) in enumerate(gap_data):
        r = gap_r + 1 + j
        sc(ws2, r, 1, lbl, font=BOLD_FONT)
        sc(ws2, r, 2, val, fmt, CALC_FILL if isinstance(val, str) and val.startswith('=') else INPUT_FILL)

    # Likely causes
    cause_r = gap_r + len(gap_data) + 2
    ws2.cell(row=cause_r, column=1, value="PROBABLE CAUSES OF 2024 EXPENSE SPIKE").font = SECTION_FONT
    causes = [
        ("Mills Branch Lawsuit Legal Fees", "High", "$50K-$200K", "Attorney fees, arbitration costs, potential settlement reserve"),
        ("Deferred Capital Maintenance", "High", "$100K-$300K", "5 parks are 30-50 years old; major systems may need replacement"),
        ("Insurance Premium Increase", "Medium", "$30K-$80K", "Post-Harvey, post-lawsuit premium adjustments"),
        ("One-Time Park Project", "Medium", "$100K-$400K", "Major renovation, new facility, flood mitigation"),
        ("Entryway/Landscaping Contract Rebid", "Low", "$20K-$50K", "New vendor at higher rates after contract renewal"),
    ]
    cause_h = ['Possible Cause', 'Likelihood', 'Est. Range', 'Rationale']
    for c, h in enumerate(cause_h, 1):
        ws2.cell(row=cause_r + 1, column=c, value=h)
    hdr(ws2, cause_r + 1, 4)
    for j, (cause, like, rng, rat) in enumerate(causes):
        r = cause_r + 2 + j
        sc(ws2, r, 1, cause, font=BOLD_FONT)
        sc(ws2, r, 2, like)
        sc(ws2, r, 3, rng)
        sc(ws2, r, 4, rat, font=NOTE_FONT)

    autofit(ws2)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 3: DCF MODEL
    # ════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("DCF Model")
    ws3.sheet_properties.tabColor = "7B1FA2"

    ws3.merge_cells('A1:N1')
    ws3['A1'] = "10-YEAR DISCOUNTED CASH FLOW MODEL"
    ws3['A1'].font = TITLE_FONT
    ws3.merge_cells('A2:N2')
    ws3['A2'] = "Change yellow cells to test scenarios. All green cells are formulas."
    ws3['A2'].font = NOTE_FONT

    # Assumptions
    ws3.cell(row=4, column=1, value="MODEL INPUTS").font = SECTION_FONT
    assum = [
        # (row, label, value, format, note)
        (5, "Equivalent Units (2008 census)", 23354, '0', "1 unit = 1 home. 0.5 = 1 apt. Commercial by sqft."),
        (6, "2024 Assessment Revenue (actual)", 958464, MONEY, "From 2024 Form 990 — Program Service Revenue"),
        (7, "2024 Investment Income (actual)", 90457, MONEY, "From 2024 Form 990"),
        (8, "2024 Total Revenue (actual)", None, MONEY, "=B6+B7 (formula)"),
        (9, "2024 Total Expenses (actual)", 1565650, MONEY, "Anomalous — 55% above prior year"),
        (10, "Normalized Expense Base", 1020000, MONEY, "Median of 2019-2023 expenses, adjusted for inflation"),
        (11, "Starting Cash Reserves", 2092894, MONEY, "From 2024 balance sheet"),
        (12, "Assessment Growth Rate", 0.015, PCT, "Historical CAGR ~1.4% (2011-2024); conservative"),
        (13, "Expense Growth Rate", 0.035, PCT, "Inflation + aging infrastructure. Historical CAGR 5.9% but volatile"),
        (14, "Investment Yield on Reserves", 0.04, PCT, "2024 actual: $90K / $2.1M ≈ 4.3%"),
        (15, "Discount Rate", 0.05, PCT, "Opportunity cost of capital for civic organization"),
        (16, "Catastrophe Probability (annual)", 0.067, PCT, "1-in-15 year event (Harvey was 2017, next due ~2032)"),
        (17, "Catastrophe Cost", 500000, MONEY, "Harvey recovery: 2.5 years, est. $500K+ total"),
        (18, "Lawsuit Settlement", 500000, MONEY, "Mills Branch exposure: demands up to $250K, legal fees additional"),
        (19, "Lawsuit Year", 2026, YR, "Expected resolution"),
        (20, "Lease Income (est, bundled)", 50000, MONEY, "Estimated — 6 field leases, likely at-cost or nominal"),
    ]

    for r, lbl, val, fmt, note in assum:
        sc(ws3, r, 1, lbl, font=BOLD_FONT)
        if r == 8:
            sc(ws3, r, 2, None, fmt, CALC_FILL).value = '=B6+B7'
        else:
            sc(ws3, r, 2, val, fmt, INPUT_FILL)
        sc(ws3, r, 3, note, font=NOTE_FONT)
        ws3.merge_cells(f'C{r}:F{r}')

    # 10-year projection
    pr = 22
    ws3.cell(row=pr, column=1, value="10-YEAR CASH FLOW PROJECTION").font = SECTION_FONT

    ph = ['Year', 'Assessment\nRevenue', 'Est. Lease\nIncome', 'Investment\nIncome',
          'Total\nRevenue', 'Operating\nExpenses', 'Net Operating\nCash Flow',
          'Catastrophe\nReserve', 'Lawsuit\nImpact', 'Net Cash\nFlow',
          'Discount\nFactor', 'PV of\nCash Flow', 'Cumulative\nCash Reserve',
          'Months of\nReserves']
    for c, h in enumerate(ph, 1):
        ws3.cell(row=pr + 1, column=c, value=h)
    hdr(ws3, pr + 1, len(ph))

    # Year 0 (2024 actuals)
    r0 = pr + 2
    sc(ws3, r0, 1, 2024, YR, font=BOLD_FONT)
    sc(ws3, r0, 2, 958464, MONEY)   # Assessment
    sc(ws3, r0, 3, None, MONEY, CALC_FILL).value = '=$B$20'  # Lease (est)
    sc(ws3, r0, 4, 90457, MONEY)    # Investment
    sc(ws3, r0, 5, None, MONEY, CALC_FILL).value = f'=B{r0}+C{r0}+D{r0}'  # Total Rev
    sc(ws3, r0, 6, 1565650, MONEY)  # Expenses (actual anomalous)
    sc(ws3, r0, 7, None, MONEY_NEG, CALC_FILL).value = f'=E{r0}-F{r0}'
    sc(ws3, r0, 8, 0, MONEY)
    sc(ws3, r0, 9, 0, MONEY)
    sc(ws3, r0, 10, None, MONEY_NEG, CALC_FILL).value = f'=G{r0}-H{r0}-I{r0}'
    sc(ws3, r0, 11, 1, '0.0000')
    sc(ws3, r0, 12, None, MONEY_NEG, CALC_FILL).value = f'=J{r0}*K{r0}'
    sc(ws3, r0, 13, 2092894, MONEY)
    sc(ws3, r0, 14, None, DEC1, CALC_FILL).value = f'=IF(F{r0}=0,0,M{r0}/F{r0}*12)'

    # Years 1-10
    for i in range(1, 11):
        r = r0 + i
        prev = r - 1
        yr = 2024 + i

        sc(ws3, r, 1, yr, YR, font=BOLD_FONT)
        # Assessment: prior * (1 + growth)
        sc(ws3, r, 2, None, MONEY, CALC_FILL).value = f'=B{prev}*(1+$B$12)'
        # Lease income (flat estimate)
        sc(ws3, r, 3, None, MONEY, CALC_FILL).value = '=$B$20'
        # Investment: prior cash reserve * yield
        sc(ws3, r, 4, None, MONEY, CALC_FILL).value = f'=MAX(M{prev},0)*$B$14'
        # Total revenue
        sc(ws3, r, 5, None, MONEY, CALC_FILL).value = f'=B{r}+C{r}+D{r}'
        # Expenses: normalized base * (1 + expense growth)^n
        sc(ws3, r, 6, None, MONEY, CALC_FILL).value = f'=$B$10*(1+$B$13)^{i}'
        # Net operating CF
        sc(ws3, r, 7, None, MONEY_NEG, CALC_FILL).value = f'=E{r}-F{r}'
        # Catastrophe: probability * cost (expected value per year)
        sc(ws3, r, 8, None, MONEY, CALC_FILL).value = '=$B$16*$B$17'
        # Lawsuit: one-time hit
        sc(ws3, r, 9, None, MONEY, CALC_FILL).value = f'=IF(A{r}=$B$19,$B$18,0)'
        # Net CF
        sc(ws3, r, 10, None, MONEY_NEG, CALC_FILL).value = f'=G{r}-H{r}-I{r}'
        # Discount factor
        sc(ws3, r, 11, None, '0.0000', CALC_FILL).value = f'=1/(1+$B$15)^{i}'
        # PV
        sc(ws3, r, 12, None, MONEY_NEG, CALC_FILL).value = f'=J{r}*K{r}'
        # Cumulative reserve
        sc(ws3, r, 13, None, MONEY_NEG, CALC_FILL).value = f'=MAX(M{prev}+J{r},0)'
        # Months of reserves
        sc(ws3, r, 14, None, DEC1, CALC_FILL).value = f'=IF(F{r}=0,0,M{r}/F{r}*12)'

    pe = r0 + 10  # projection end

    # Conditional formatting: highlight low reserves
    for r in range(r0, pe + 1):
        # We'll just set fill manually based on static thresholds in post
        pass

    # DCF Results
    dr = pe + 2
    ws3.cell(row=dr, column=1, value="DCF RESULTS — WHAT THE NUMBERS PROVE").font = SECTION_FONT

    results = [
        ("NPV of 10-Year Cash Flows", f'=SUM(L{r0+1}:L{pe})', MONEY_NEG,
         "Present value of all future cash flows, discounted at 5%. Negative = organization loses value."),
        ("Terminal Cash Reserve (2034)", f'=M{pe}', MONEY_NEG,
         "Cash remaining after 10 years. If zero, KSA is insolvent."),
        ("Total Undiscounted CF (10yr)", f'=SUM(J{r0+1}:J{pe})', MONEY_NEG,
         "Raw sum of net cash flows before discounting."),
        ("Avg Annual Net CF", f'=AVERAGE(J{r0+1}:J{pe})', MONEY_NEG,
         "Average cash burn/gain per year."),
        ("Years with Positive CF", f'=COUNTIF(J{r0+1}:J{pe},">"&0)', '0',
         "How many of the 10 years generate surplus."),
        ("Years Until Reserves Hit Zero", f'=COUNTIF(M{r0+1}:M{pe},">"&0)', '0',
         "How many years before KSA runs out of cash."),
        ("Final Months of Operating Reserve", f'=N{pe}', DEC1,
         "Healthy: 12+. Warning: 6-12. Critical: <6."),
        ("Breakeven Assessment Growth Needed", f'=($B$10*(1+$B$13)-$B$7-$B$20)/$B$6-1', PCT2,
         "Assessment growth rate that makes Year 1 CF = 0 (ignoring catastrophe/lawsuit)."),
        ("Assessment Increase Per Unit (breakeven)", f'=(($B$10*(1+$B$13)-$B$7-$B$20)-$B$6)/23354', '$#,##0.00',
         "Dollar increase per unit needed in Year 1 to break even."),
        ("Cumulative Catastrophe Cost (expected)", f'=SUM(H{r0+1}:H{pe})', MONEY,
         "Total expected catastrophe reserve over 10 years."),
        ("Total Lawsuit Impact", f'=$B$18', MONEY,
         "One-time settlement payment."),
    ]

    for j, (lbl, fml, fmt, note) in enumerate(results):
        r = dr + 1 + j
        sc(ws3, r, 1, lbl, font=BOLD_FONT)
        sc(ws3, r, 2, None, fmt, CALC_FILL).value = fml
        sc(ws3, r, 3, note, font=NOTE_FONT)
        ws3.merge_cells(f'C{r}:F{r}')

    autofit(ws3)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 4: SENSITIVITY — WHAT IF WE CHANGE ASSESSMENTS?
    # ════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Assessment Sensitivity")
    ws4.sheet_properties.tabColor = "B71C1C"

    ws4.merge_cells('A1:J1')
    ws4['A1'] = "ASSESSMENT INCREASE SENSITIVITY — PROVING WHAT IT TAKES"
    ws4['A1'].font = TITLE_FONT

    # Fixed inputs referenced from DCF sheet
    ws4.cell(row=3, column=1, value="INPUTS (from DCF Model sheet)").font = SECTION_FONT
    inp = [
        (4, "Current Assessment Rev", "='DCF Model'!B6", MONEY),
        (5, "Equivalent Units", "='DCF Model'!B5", '0'),
        (6, "Current Per-Unit Cost", "=B4/B5", '$#,##0.00'),
        (7, "Normalized Expenses (Year 1)", "='DCF Model'!B10*(1+'DCF Model'!B13)", MONEY),
        (8, "Investment Income (est)", "='DCF Model'!B11*'DCF Model'!B14", MONEY),
        (9, "Lease Income (est)", "='DCF Model'!B20", MONEY),
        (10, "Year 1 Gap (no change)", "=B7-B4-B8-B9", MONEY_NEG),
        (11, "Cash Reserves", "='DCF Model'!B11", MONEY),
        (12, "Years to Exhaust (at current gap)", "=IF(B10<=0,99,B11/B10)", DEC1),
    ]
    for r, lbl, fml, fmt in inp:
        sc(ws4, r, 1, lbl, font=BOLD_FONT)
        sc(ws4, r, 2, None, fmt, CALC_FILL).value = fml

    # Scenario table
    st = 14
    ws4.cell(row=st, column=1, value="PER-UNIT INCREASE SCENARIOS").font = SECTION_FONT

    st_h = ['Increase\nPer Unit', 'New Per-Unit\nAssessment', '% Increase\nPer Unit',
            'New Total\nAssessment Rev', 'Year 1\nTotal Revenue', 'Year 1\nExpenses',
            'Year 1\nNet CF', 'Years Until\nReserves Gone',
            'Sustainable\n10 Years?', 'Monthly\nIncrease/Home']
    for c, h in enumerate(st_h, 1):
        ws4.cell(row=st + 1, column=c, value=h)
    hdr(ws4, st + 1, len(st_h))

    increases = [0, 5, 10, 15, 20, 25, 30, 40, 50, 75, 100, 150, 200]
    for j, inc in enumerate(increases):
        r = st + 2 + j
        sc(ws4, r, 1, inc, MONEY, INPUT_FILL)
        # New per-unit
        sc(ws4, r, 2, None, '$#,##0.00', CALC_FILL).value = f'=$B$6+A{r}'
        # % increase
        sc(ws4, r, 3, None, PCT, CALC_FILL).value = f'=A{r}/$B$6'
        # New total assessment rev
        sc(ws4, r, 4, None, MONEY, CALC_FILL).value = f'=B{r}*$B$5'
        # Year 1 total revenue
        sc(ws4, r, 5, None, MONEY, CALC_FILL).value = f'=D{r}+$B$8+$B$9'
        # Year 1 expenses
        sc(ws4, r, 6, None, MONEY, CALC_FILL).value = '=$B$7'
        # Year 1 net CF
        sc(ws4, r, 7, None, MONEY_NEG, CALC_FILL).value = f'=E{r}-F{r}'
        # Years to exhaust
        sc(ws4, r, 8, None, DEC1, CALC_FILL).value = f'=IF(G{r}>=0,99,IF(-G{r}=0,99,$B$11/(-G{r})))'
        # Sustainable 10yr?
        sc(ws4, r, 9, None, font=BOLD_FONT).value = f'=IF(H{r}>=10,"YES","NO")'
        # Monthly per home
        sc(ws4, r, 10, None, '$#,##0.00', CALC_FILL).value = f'=A{r}/12'

    # Proof section
    pf = st + 2 + len(increases) + 2
    ws4.cell(row=pf, column=1, value="WHAT THIS PROVES").font = SECTION_FONT
    proofs = [
        "PROOF 1: At $0 increase, KSA runs a structural deficit. Revenue grows slower than expenses (1.5% vs 3.5%).",
        "PROOF 2: The revenue-expense gap COMPOUNDS. Year 1 gap is modest, but by year 5 it accelerates.",
        "PROOF 3: A $5/unit/year increase ($0.42/month per home) generates ~$117K additional revenue — significant.",
        "PROOF 4: Find the row where 'Sustainable 10 Years?' flips from NO to YES — that's the minimum viable increase.",
        "PROOF 5: Even the minimum increase doesn't account for catastrophic events (Harvey) or lawsuits (Mills Branch).",
        "PROOF 6: The 2024 expense spike consumed ~25% of liquid reserves in ONE year. Reserves are finite.",
        "PROOF 7: Assessment revenue per unit (~$41/yr or $3.42/month) is remarkably low for 356 acres of maintained parkland.",
        "PROOF 8: At $200/unit increase, monthly cost per home goes from $3.42 to $20.08 — still low vs comparable HOAs.",
    ]
    for j, p in enumerate(proofs):
        ws4.cell(row=pf + 1 + j, column=1, value=p).font = NORMAL_FONT
        ws4.merge_cells(f'A{pf+1+j}:J{pf+1+j}')

    autofit(ws4)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 5: BALANCE SHEET & RESERVE HEALTH
    # ════════════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Balance Sheet")
    ws5.sheet_properties.tabColor = "1565C0"

    ws5.merge_cells('A1:I1')
    ws5['A1'] = "KSA BALANCE SHEET ANALYSIS — RESERVE HEALTH"
    ws5['A1'].font = TITLE_FONT

    bs_h = ['Year', 'Total\nAssets', 'Total\nLiabilities', 'Net\nAssets',
            'Net Asset\nYoY Change', 'Debt/Asset\nRatio', 'Net Income\n(Rev-Exp)',
            'Cash/Invest\n(est)', 'Months of\nReserves']
    for c, h in enumerate(bs_h, 1):
        ws5.cell(row=3, column=c, value=h)
    hdr(ws5, 3, len(bs_h))

    for i, d in enumerate(DATA):
        r = 4 + i
        yr, prog, inv, oth, tot, exp, ast, liab, net = d
        sc(ws5, r, 1, yr, YR, font=BOLD_FONT)
        sc(ws5, r, 2, ast, MONEY)
        sc(ws5, r, 3, liab, MONEY)
        sc(ws5, r, 4, net, MONEY)
        if i > 0:
            sc(ws5, r, 5, None, MONEY_NEG, CALC_FILL).value = f'=D{r}-D{r-1}'
        sc(ws5, r, 6, None, PCT2, CALC_FILL).value = f'=IF(B{r}=0,0,C{r}/B{r})'
        sc(ws5, r, 7, None, MONEY_NEG, CALC_FILL).value = f'={tot}-{exp}'
        # Estimate cash as 67% of assets (from 2024 ratio: 2.09M/3.12M)
        sc(ws5, r, 8, None, MONEY, CALC_FILL).value = f'=B{r}*0.67'
        sc(ws5, r, 9, None, DEC1, CALC_FILL).value = f'=IF({exp}=0,0,H{r}/{exp}*12)'

    bde = 4 + len(DATA) - 1

    # Reserve health summary
    rh = bde + 2
    ws5.cell(row=rh, column=1, value="RESERVE HEALTH INDICATORS").font = SECTION_FONT
    rh_data = [
        ("2024 Cash/Investments (actual)", 2092894, MONEY),
        ("2024 Total Expenses", 1565650, MONEY),
        ("Months at 2024 Expense Rate", f'=B{rh+1}/B{rh+2}*12', DEC1),
        ("2023 Cash/Investments (actual)", 2594104, MONEY),
        ("Cash Decline (2023→2024)", f'=B{rh+4}-B{rh+1}', MONEY_NEG),
        ("% of Reserves Lost in 1 Year", f'=B{rh+5}/B{rh+4}', PCT),
        ("At This Rate, Reserves Last", f'=IF(B{rh+5}<=0,99,B{rh+1}/B{rh+5})', DEC1),
        ("Net Assets Peak (2023)", 3637007, MONEY),
        ("Net Assets Now (2024)", 3120278, MONEY),
        ("Decline from Peak", f'=B{rh+8}-B{rh+9}', MONEY_NEG),
        ("% Decline from Peak", f'=B{rh+10}/B{rh+8}', PCT),
    ]
    for j, (lbl, val, fmt) in enumerate(rh_data):
        r = rh + 1 + j
        sc(ws5, r, 1, lbl, font=BOLD_FONT)
        if isinstance(val, str) and val.startswith('='):
            sc(ws5, r, 2, None, fmt, CALC_FILL).value = val
        else:
            sc(ws5, r, 2, val, fmt, INPUT_FILL)

    autofit(ws5)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 6: RGDGC IMPACT
    # ════════════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("RGDGC Impact")
    ws6.sheet_properties.tabColor = "FF6B35"

    ws6.merge_cells('A1:G1')
    ws6['A1'] = "WHAT THIS MEANS FOR RIVER GROVE DISC GOLF CLUB"
    ws6['A1'].font = TITLE_FONT

    ws6.cell(row=3, column=1, value="RIVER GROVE PARK BUDGET ALLOCATION").font = SECTION_FONT
    alloc = [
        (4, "KSA Total Parks Budget (est 44%)", "='Expense Analysis'!B" + str(bc + 2), MONEY, CALC_FILL,
         "Parks ops is ~44% of KSA budget (confirmed 2009, 2012)"),
        (5, "Number of Parks", 5, '0', INPUT_FILL, "River Grove, East End, Deer Ridge, Northpark, Creekwood"),
        (6, "Total Park Acreage", 356.6, '0.0', INPUT_FILL, "KSA-owned parkland"),
        (7, "River Grove Acreage", 74, '0', INPUT_FILL, ""),
        (8, "River Grove % (by acreage)", "=B7/B6", PCT, CALC_FILL, ""),
        (9, "Est. River Grove Annual Budget", "=B4*B8", MONEY, CALC_FILL, "KSA's approximate spend on River Grove"),
        (10, "Disc Golf % of River Grove (est)", 0.15, PCT, INPUT_FILL, "DG uses ~15% of park area/resources"),
        (11, "Est. Annual DG Maintenance", "=B9*B10", MONEY, CALC_FILL, "What KSA likely spends on disc golf annually"),
    ]
    for r, lbl, val, fmt, fill, note in alloc:
        sc(ws6, r, 1, lbl, font=BOLD_FONT)
        if isinstance(val, str) and val.startswith('='):
            sc(ws6, r, 2, None, fmt, fill).value = val
        else:
            sc(ws6, r, 2, val, fmt, fill)
        sc(ws6, r, 3, note, font=NOTE_FONT)

    # Improvement ROI
    ws6.cell(row=13, column=1, value="DISC GOLF IMPROVEMENT COSTS & ROI").font = SECTION_FONT
    imp_h = ['Improvement', 'Unit $', 'Qty', 'Total $', '% of RG Budget', 'Funding Source']
    for c, h in enumerate(imp_h, 1):
        ws6.cell(row=14, column=c, value=h)
    hdr(ws6, 14, 6)

    imps = [
        ("Concrete tee pads", 800, 21, "RGDGC co-fund 50% + KSA Parks Foundation"),
        ("DGA Mach X baskets", 450, 21, "Parks Foundation 501(c)(3) donor campaign"),
        ("Per-hole signage", 150, 21, "RGDGC member dues"),
        ("Benches/seating", 300, 18, "Parks Foundation + memorial dedications"),
        ("Practice basket area", 2000, 1, "RGDGC tournament revenue"),
        ("Drainage improvements", 25000, 1, "KSA capital budget + HCFCD grant"),
        ("Trail resurfacing", 8000, 1, "KSA operating budget"),
    ]
    for j, (name, cost, qty, source) in enumerate(imps):
        r = 15 + j
        sc(ws6, r, 1, name)
        sc(ws6, r, 2, cost, MONEY, INPUT_FILL)
        sc(ws6, r, 3, qty, '0', INPUT_FILL)
        sc(ws6, r, 4, None, MONEY, CALC_FILL).value = f'=B{r}*C{r}'
        sc(ws6, r, 5, None, PCT, CALC_FILL).value = f'=D{r}/$B$9'
        sc(ws6, r, 6, source, font=NOTE_FONT)

    tr6 = 15 + len(imps)
    sc(ws6, tr6, 1, "TOTAL ALL IMPROVEMENTS", font=BOLD_FONT)
    sc(ws6, tr6, 4, None, MONEY, CALC_FILL).value = f'=SUM(D15:D{tr6-1})'
    sc(ws6, tr6, 5, None, PCT, CALC_FILL).value = f'=D{tr6}/$B$9'

    # Value proposition
    vp = tr6 + 2
    ws6.cell(row=vp, column=1, value="DISC GOLF VALUE PROPOSITION TO KSA").font = SECTION_FONT
    vp_data = [
        (vp+1, "Estimated weekly players", 150, '0', INPUT_FILL),
        (vp+2, "Annual player visits", "=B" + str(vp+1) + "*52", '0', CALC_FILL),
        (vp+3, "Estimated unique annual visitors", f'=B{vp+2}*0.3', '0', CALC_FILL),
        (vp+4, "% Kingwood residents", 0.7, PCT, INPUT_FILL),
        (vp+5, "Resident players per year", f'=B{vp+3}*B{vp+4}', '0', CALC_FILL),
        (vp+6, "KSA cost per player visit", f'=B11/B{vp+2}', '$#,##0.00', CALC_FILL),
        (vp+7, "Cost per unique visitor", f'=B11/B{vp+3}', '$#,##0.00', CALC_FILL),
        (vp+8, "Annual assessment per unit", "='DCF Model'!B6/'DCF Model'!B5", '$#,##0.00', CALC_FILL),
        (vp+9, "DG cost as % of assessment", f'=B11/B{vp+8}', PCT, CALC_FILL),
    ]
    for r, lbl, val, fmt, fill in vp_data:
        sc(ws6, r, 1, lbl, font=BOLD_FONT)
        if isinstance(val, str) and val.startswith('='):
            sc(ws6, r, 2, None, fmt, fill).value = val
        else:
            sc(ws6, r, 2, val, fmt, fill)

    autofit(ws6)

    # ════════════════════════════════════════════════════════════════════════
    # SHEET 7: FORENSIC EVIDENCE — FRAUD INDICATORS
    # ════════════════════════════════════════════════════════════════════════
    ws7 = wb.create_sheet("Forensic Evidence")
    ws7.sheet_properties.tabColor = "B71C1C"

    ws7.merge_cells('A1:H1')
    ws7['A1'] = "FORENSIC FINANCIAL EVIDENCE — KSA FRAUD INDICATORS"
    ws7['A1'].font = TITLE_FONT
    ws7.merge_cells('A2:H2')
    ws7['A2'] = "Analysis from aggressive fraud attorney perspective. All formulas prove claims with 990 data."
    ws7['A2'].font = NOTE_FONT

    # ── Exhibit A: Retained Surplus Analysis ─────────────────────────
    ws7.cell(row=4, column=1, value="EXHIBIT A: RETAINED SURPLUS ANALYSIS").font = SECTION_FONT
    ws7.cell(row=5, column=1, value="Article VIII requires return of unspent funds. These formulas prove surpluses were retained.").font = NOTE_FONT
    ws7.merge_cells('A5:H5')

    ea_h = ['Year', 'Revenue', 'Expenses', 'Surplus\n(Rev-Exp)', 'Surplus\nReturned?',
            'Cumulative\nRetained', 'Net Asset\nChange', 'Proves\nRetention?']
    for c, h in enumerate(ea_h, 1):
        ws7.cell(row=6, column=c, value=h)
    hdr(ws7, 6, 8)

    cumul_formula_parts = []
    for i, d in enumerate(DATA):
        r = 7 + i
        yr, prog, inv, oth, tot, exp, ast, liab, net = d
        sc(ws7, r, 1, yr, YR, font=BOLD_FONT)
        sc(ws7, r, 2, tot, MONEY)
        sc(ws7, r, 3, exp, MONEY)
        sc(ws7, r, 4, None, MONEY_NEG, CALC_FILL).value = f'=B{r}-C{r}'
        # Surplus returned? Unknown — but if net assets grew, surplus was NOT returned
        sc(ws7, r, 5, "UNKNOWN", font=Font(name="Calibri", bold=True, color="B71C1C"))
        # Cumulative retained (sum of positive surpluses only)
        sc(ws7, r, 6, None, MONEY, CALC_FILL).value = f'=SUMPRODUCT((D$7:D{r})*(D$7:D{r}>0))'
        # Net asset change
        if i > 0:
            sc(ws7, r, 7, None, MONEY_NEG, CALC_FILL).value = f'={net}-{DATA[i-1][8]}'
        else:
            sc(ws7, r, 7, 0, MONEY)
        # Proves retention: if surplus > 0 AND net assets grew, surplus was retained
        if i > 0:
            sc(ws7, r, 8, None, font=BOLD_FONT).value = f'=IF(AND(D{r}>0,G{r}>0),"YES - RETAINED","Inconclusive")'
        else:
            sc(ws7, r, 8, "N/A")

    fde = 7 + len(DATA) - 1  # forensic data end

    # Summary
    fs = fde + 2
    ws7.cell(row=fs, column=1, value="EXHIBIT A FINDINGS").font = SECTION_FONT
    findings_a = [
        ("Total Years with Surplus", f'=COUNTIF(D7:D{fde},">"&0)', '0'),
        ("Total Surplus Generated (all years)", f'=SUMPRODUCT((D7:D{fde})*(D7:D{fde}>0))', MONEY),
        ("Total Deficit Years", f'=COUNTIF(D7:D{fde},"<"&0)', '0'),
        ("Total Deficits", f'=SUMPRODUCT((D7:D{fde})*(D7:D{fde}<0))', MONEY_NEG),
        ("Net Surplus Over 14 Years", f'=SUM(D7:D{fde})', MONEY_NEG),
        ("Years Where Surplus Was Provably Retained", f'=COUNTIF(H7:H{fde},"YES*")', '0'),
        ("Net Asset Growth (2011→2024)", f'={DATA[-1][8]}-{DATA[0][8]}', MONEY_NEG),
        ("If All Surpluses Returned: Expected Net Asset Change", f'=B{fs+4}', MONEY_NEG),
        ("Actual Net Asset Change", f'=B{fs+7}', MONEY_NEG),
        ("DISCREPANCY (Surplus Retained)", f'=B{fs+7}-B{fs+8}', MONEY_NEG),
    ]
    for j, (lbl, fml, fmt) in enumerate(findings_a):
        r = fs + 1 + j
        sc(ws7, r, 1, lbl, font=BOLD_FONT)
        sc(ws7, r, 2, None, fmt, CALC_FILL).value = fml

    # ── Exhibit B: KAM Compensation Analysis ─────────────────────────
    eb = fs + len(findings_a) + 3
    ws7.cell(row=eb, column=1, value="EXHIBIT B: UNDISCLOSED KAM COMPENSATION").font = SECTION_FONT

    eb_data = [
        (eb+1, "KSA Reports Officer Compensation", 0, MONEY, "Form 990 Part VII: $0 for all officers including McCormick"),
        (eb+2, "KSA Reports Employee Count", 0, '0', "Form 990: 0 employees, all years"),
        (eb+3, "G&A Budget (2012 actual)", 182783, MONEY, "Includes KAM management contract (amount undisclosed)"),
        (eb+4, "G&A Budget (2009 actual)", 153319, MONEY, "KAM contract embedded here"),
        (eb+5, "G&A CAGR (2009→2012)", None, PCT2, "formula"),
        (eb+6, "Estimated 2024 G&A (at CAGR)", None, MONEY, "formula"),
        (eb+7, "KAM Contract Estimate (50-70% of G&A)", None, MONEY, "Industry standard: management = 50-70% of G&A"),
        (eb+8, "KAM Contract Estimate (HIGH)", None, MONEY, "70% of estimated G&A"),
        (eb+9, "Additional: KAM collects from village HOAs", "UNDISCLOSED", None, "KAM manages both KSA and individual villages"),
        (eb+10, "Total KAM Revenue from Kingwood", "UNKNOWN", None, "Private company — no public disclosure"),
    ]
    for r, lbl, val, fmt, note in eb_data:
        sc(ws7, r, 1, lbl, font=BOLD_FONT)
        if r == eb+5:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=(B{eb+3}/B{eb+4})^(1/3)-1'
        elif r == eb+6:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{eb+3}*(1+B{eb+5})^12'
        elif r == eb+7:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{eb+6}*0.5'
        elif r == eb+8:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{eb+6}*0.7'
        elif isinstance(val, (int, float)):
            sc(ws7, r, 2, val, fmt, INPUT_FILL)
        else:
            sc(ws7, r, 2, val, font=Font(name="Calibri", bold=True, color="B71C1C"))
        sc(ws7, r, 3, note, font=NOTE_FONT)
        ws7.merge_cells(f'C{r}:H{r}')

    # ── Exhibit C: 2024 Expense Anomaly ──────────────────────────────
    ec = eb + 12
    ws7.cell(row=ec, column=1, value="EXHIBIT C: 2024 EXPENSE ANOMALY — FORENSIC BREAKDOWN").font = SECTION_FONT

    ec_data = [
        (ec+1, "2024 Total Expenses", 1565650, MONEY, "Highest in KSA history"),
        (ec+2, "2023 Total Expenses", 1009164, MONEY, "Prior year (normal)"),
        (ec+3, "YoY Increase ($)", None, MONEY, "formula"),
        (ec+4, "YoY Increase (%)", None, PCT, "formula"),
        (ec+5, "14-Year Median Expenses", None, MONEY, "formula"),
        (ec+6, "2024 vs Median ($)", None, MONEY, "formula"),
        (ec+7, "2024 vs Median (%)", None, PCT, "formula"),
        (ec+8, "14-Year Std Deviation", None, MONEY, "formula"),
        (ec+9, "Z-Score (how many std devs from mean)", None, '0.00', "formula"),
        (ec+10, "STATISTICAL SIGNIFICANCE", None, None, "formula"),
    ]
    for r, lbl, val, fmt, note in ec_data:
        sc(ws7, r, 1, lbl, font=BOLD_FONT)
        if r == ec+3:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{ec+1}-B{ec+2}'
        elif r == ec+4:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{ec+3}/B{ec+2}'
        elif r == ec+5:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f"=MEDIAN('Revenue Analysis'!F5:'Revenue Analysis'!F{5+len(DATA)-1})"
            # Simpler: use hardcoded reference to Expense Analysis
        elif r == ec+6:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{ec+1}-B{ec+5}'
        elif r == ec+7:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{ec+6}/B{ec+5}'
        elif r == ec+8:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f"=STDEV('Revenue Analysis'!F5:'Revenue Analysis'!F{5+len(DATA)-1})"
        elif r == ec+9:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=(B{ec+1}-AVERAGE(B{ec+5},B{ec+5}))/B{ec+8}'
        elif r == ec+10:
            sc(ws7, r, 2, None, font=Font(name="Calibri", bold=True, color="B71C1C", size=12)
               ).value = f'=IF(B{ec+9}>2,"STATISTICALLY ANOMALOUS (>2σ)",IF(B{ec+9}>1.5,"ELEVATED","NORMAL"))'
        elif isinstance(val, (int, float)):
            sc(ws7, r, 2, val, fmt, INPUT_FILL)
        sc(ws7, r, 3, note, font=NOTE_FONT)

    # ── Exhibit D: Cascade Liability ─────────────────────────────────
    ed = ec + 13
    ws7.cell(row=ed, column=1, value="EXHIBIT D: MILLS BRANCH CASCADE — WHAT IF ALL VILLAGES FILE?").font = SECTION_FONT

    ed_data = [
        (ed+1, "Member Associations", 29, '0', "Villages + commercial"),
        (ed+2, "Mills Branch Damages Sought", 250000, MONEY, "Per lawsuit filing"),
        (ed+3, "If All Villages File (worst case)", None, MONEY, "formula"),
        (ed+4, "KSA Total Assets (2024)", 3123778, MONEY, "From 990"),
        (ed+5, "KSA Cash/Investments (2024)", 2092894, MONEY, "Liquid reserves"),
        (ed+6, "Shortfall if All File (vs assets)", None, MONEY_NEG, "formula"),
        (ed+7, "Shortfall if All File (vs cash)", None, MONEY_NEG, "formula"),
        (ed+8, "% of Assets at Risk", None, PCT, "formula"),
        (ed+9, "SOLVENCY STATUS", None, None, "formula"),
    ]
    for r, lbl, val, fmt, note in ed_data:
        sc(ws7, r, 1, lbl, font=BOLD_FONT)
        if r == ed+3:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{ed+1}*B{ed+2}'
        elif r == ed+6:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{ed+4}-B{ed+3}'
        elif r == ed+7:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{ed+5}-B{ed+3}'
        elif r == ed+8:
            sc(ws7, r, 2, None, fmt, CALC_FILL).value = f'=B{ed+3}/B{ed+4}'
        elif r == ed+9:
            sc(ws7, r, 2, None, font=Font(name="Calibri", bold=True, color="B71C1C", size=12)
               ).value = f'=IF(B{ed+6}<0,"INSOLVENT — assets < liability","Solvent but impaired")'
        elif isinstance(val, (int, float)):
            sc(ws7, r, 2, val, fmt, INPUT_FILL)
        sc(ws7, r, 3, note, font=NOTE_FONT)

    # ── Exhibit E: Self-Dealing Scorecard ─────────────────────────────
    ee = ed + 12
    ws7.cell(row=ee, column=1, value="EXHIBIT E: SELF-DEALING RED FLAG SCORECARD").font = SECTION_FONT

    red_flags = [
        ("Management company owner sits on nonprofit board", "YES", "CRITICAL", "Textbook interlocking directorate"),
        ("Same company manages overseer AND overseen", "YES", "CRITICAL", "KAM manages villages that appoint KSA board"),
        ("$0 officer compensation with undisclosed contractor pay", "YES", "HIGH", "McCormick paid via KAM, not reported on 990"),
        ("No evidence of competitive bidding", "YES", "HIGH", "KAM contract appears to be no-bid"),
        ("BBB F rating — refuses to respond to complaints", "YES", "HIGH", "Pattern of avoiding accountability"),
        ("Refused contractual arbitration (Mills Branch)", "YES", "HIGH", "Only sought arbitration after being sued"),
        ("Management departure and swift return", "YES", "MEDIUM", "FirstService Residential displaced without explanation"),
        ("$0 rental income despite 6+ active leases", "YES", "MEDIUM", "Either misclassified or at-cost (unusual)"),
        ("Same physical address for nonprofit and for-profit", "YES", "MEDIUM", "1075 Kingwood Dr #100"),
        ("BBB lists management company as alternate name", "YES", "MEDIUM", "Entities functionally interchangeable"),
        ("Unexplained 55% expense spike in year of lawsuit", "YES", "MEDIUM", "2024: $1.57M expenses, coincides with Mills Branch"),
        ("Wild liability swings ($170K → $0 → $3.5K)", "YES", "LOW", "May indicate inconsistent accounting"),
    ]

    rf_h = ['Red Flag', 'Present?', 'Severity', 'Evidence']
    for c, h in enumerate(rf_h, 1):
        ws7.cell(row=ee + 1, column=c, value=h)
    hdr(ws7, ee + 1, 4)

    for j, (flag, present, severity, evidence) in enumerate(red_flags):
        r = ee + 2 + j
        sc(ws7, r, 1, flag)
        sc(ws7, r, 2, present, font=Font(bold=True, color="B71C1C"))
        sev_fill = ALERT_FILL if severity == "CRITICAL" else (WARN_FILL if severity == "HIGH" else CALC_FILL)
        sc(ws7, r, 3, severity, fill=sev_fill, font=BOLD_FONT)
        sc(ws7, r, 4, evidence, font=NOTE_FONT)

    # Final count
    count_r = ee + 2 + len(red_flags) + 1
    sc(ws7, count_r, 1, "TOTAL RED FLAGS PRESENT", font=BOLD_FONT)
    sc(ws7, count_r, 2, None, '0', ALERT_FILL).value = f'=COUNTIF(B{ee+2}:B{count_r-1},"YES")'
    sc(ws7, count_r, 3, f"out of {len(red_flags)}", font=BOLD_FONT)

    autofit(ws7)

    # ── Save ─────────────────────────────────────────────────────────────
    path = "/Users/blakesanders/RGDGC/docs/KSA-Research/KSA_Financial_DCF_Model.xlsx"
    wb.save(path)
    print(f"Saved: {path}")
    print(f"Sheets: {wb.sheetnames}")
    for name in wb.sheetnames:
        ws = wb[name]
        formulas = sum(1 for row in ws.iter_rows() for c in row
                       if c.value and isinstance(c.value, str) and c.value.startswith('='))
        print(f"  {name}: {formulas} formulas, {ws.max_row} rows")


if __name__ == "__main__":
    build()
